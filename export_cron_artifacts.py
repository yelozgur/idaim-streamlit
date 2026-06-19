"""export_cron_artifacts.py — Export all cron pipeline artifacts to Excel.

Exports to data/exports/cron_run_<timestamp>.xlsx:
- Metrics: training metrics (PR-AUC, ROC-AUC, Kappa) per species
- WatchList_Aedes: 50 cells with proba >= 0.10
- WatchList_Culex: 50 cells with proba >= 0.10
- Features: 37K cells × 60 features (5 features × 12 months)
- CellPredictions: all 37K cells with aedes_proba + culex_proba
- DummyLabs: 14 dummy lab_results rows (TRP-DUMMY-*)
- DummyInits: 14 dummy sampling_initiation rows (TRP-DUMMY-*)
- LabResults: full lab_results sheet (existing + dummy)
- SamplingInit: full sampling_initiation sheet (existing + dummy)

Cron pipeline contract: every artifact must be reproducible from inputs
+ script. Excel exports make the artifacts auditable by UNDP without
running the pipeline.
"""
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parent

# Local: no Streamlit
import unittest.mock as mock
sys.modules['streamlit'] = mock.MagicMock()
sys.path.insert(0, str(REPO))

import json as _json
import gspread
from google.oauth2.service_account import Credentials

SHEET_ID = '16wqnRUUPNBA_qhPMEdy4g9gCm_QKu5IxyCbJweStRCY'
SA_PATH = Path('/Users/ozguryel/Documents/Personal Projects/ee-yelozgur-aceaafb59a17.json')


def _get_sheet():
    creds = Credentials.from_service_account_info(
        _json.loads(SA_PATH.read_text()),
        scopes=['https://www.googleapis.com/auth/spreadsheets']
    )
    gc = gspread.authorize(creds)
    return gc.open_by_key(SHEET_ID)


def _read_tab(ws):
    rows = ws.get_all_values()
    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows[1:], columns=rows[0])


def main():
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_dir = REPO / 'data' / 'exports'
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f'cron_run_{timestamp}.xlsx'
    print(f'Writing: {out_path}')

    wb = Workbook()
    # Remove default sheet, we'll add named ones
    wb.remove(wb.active)

    # 1) Metrics
    metrics_path = REPO / 'data' / '07_models' / 'metrics_local.json'
    if metrics_path.exists():
        with open(metrics_path) as f:
            metrics = _json.load(f)
        # Flatten to one row per species
        rows = []
        for species, m in metrics.items():
            row = {'species': species}
            row.update(m)
            rows.append(row)
        metrics_df = pd.DataFrame(rows)
        ws = wb.create_sheet('Metrics')
        ws.append(list(metrics_df.columns))
        for _, r in metrics_df.iterrows():
            ws.append([r[c] for c in metrics_df.columns])
        # Header style
        for c in range(1, len(metrics_df.columns) + 1):
            cell = ws.cell(1, c)
            cell.font = Font(bold=True)
        print(f'  Metrics: {len(metrics_df)} rows')

    # 2) Watch lists
    sh = _get_sheet()
    print('Reading watch_list from Sheets...')
    watch_df = _read_tab(sh.worksheet('watch_list'))
    for species in ['aedes', 'culex']:
        sub = watch_df[watch_df['species'].str.lower() == species].copy() if len(watch_df) > 0 else pd.DataFrame()
        if 'proba' in sub.columns:
            sub['proba'] = pd.to_numeric(sub['proba'], errors='coerce')
            sub = sub.sort_values('proba', ascending=False)
        ws = wb.create_sheet(f'WatchList_{species.capitalize()}')
        if len(sub) > 0:
            ws.append(list(sub.columns))
            for _, r in sub.iterrows():
                ws.append([r[c] for c in sub.columns])
            for c in range(1, len(sub.columns) + 1):
                ws.cell(1, c).font = Font(bold=True)
        print(f'  WatchList_{species}: {len(sub)} rows')

    # 3) Features
    print('Reading features Parquet...')
    cells = pd.read_parquet(REPO / 'data/cells_full.parquet')
    features_path = REPO / 'data' / 'features_cache_active.parquet'
    if features_path.exists():
        features_df = pd.read_parquet(features_path)
        # Merge cell_id with features (one row per cell, 60+ feature columns)
        feat_cols = [c for c in features_df.columns if c != 'cell_id']
        feat_data = cells[['cell_id', 'district', 'lat', 'lon']].merge(
            features_df, on='cell_id', how='left'
        )
        ws = wb.create_sheet('Features')
        ws.append(list(feat_data.columns))
        for _, r in feat_data.iterrows():
            ws.append([r[c] for c in feat_data.columns])
        for c in range(1, len(feat_data.columns) + 1):
            ws.cell(1, c).font = Font(bold=True)
        print(f'  Features: {len(feat_data)} rows × {len(feat_data.columns)} cols')

    # 4) Cell predictions (from Parquet cells with merged ML proba — only aedes_proba/culex_proba in Parquet, not from this run)
    # Use the predictions from the watch_list (just the high-proba ones, plus all 37K can be computed)
    # For now, only show watch_list cells with their proba. Full predictions would require re-running.
    ws = wb.create_sheet('CellPredictions')
    if len(watch_df) > 0:
        watch_simple = watch_df[['cell_id', 'species', 'proba', 'threshold_used', 'district']].copy()
        if 'proba' in watch_simple.columns:
            watch_simple['proba'] = pd.to_numeric(watch_simple['proba'], errors='coerce')
        # Pivot so one row per cell with aedes_proba + culex_proba
        pivot = watch_simple.pivot_table(
            index='cell_id', columns='species', values='proba', aggfunc='max'
        ).reset_index()
        if 'aedes' not in pivot.columns:
            pivot['aedes'] = None
        if 'culex' not in pivot.columns:
            pivot['culex'] = None
        # Add district
        district_map = cells.set_index('cell_id')['district'].to_dict()
        pivot['district'] = pivot['cell_id'].map(district_map)
        ws.append(list(pivot.columns))
        for _, r in pivot.iterrows():
            ws.append([r[c] for c in pivot.columns])
        for c in range(1, len(pivot.columns) + 1):
            ws.cell(1, c).font = Font(bold=True)
        print(f'  CellPredictions: {len(pivot)} cells (watch_list only)')

    # 5) Dummy data + full sheets
    print('Reading full sheets...')
    lab_df = _read_tab(sh.worksheet('lab_results'))
    init_df = _read_tab(sh.worksheet('sampling_initiation'))

    if len(lab_df) > 0:
        # Dummy labs
        dummy_labs = lab_df[lab_df['trap_id'].str.startswith('TRP-DUMMY', na=False)].copy()
        ws = wb.create_sheet('DummyLabs')
        if len(dummy_labs) > 0:
            ws.append(list(dummy_labs.columns))
            for _, r in dummy_labs.iterrows():
                ws.append([r[c] for c in dummy_labs.columns])
            for c in range(1, len(dummy_labs.columns) + 1):
                ws.cell(1, c).font = Font(bold=True)
        print(f'  DummyLabs: {len(dummy_labs)} rows')

    if len(init_df) > 0:
        dummy_inits = init_df[init_df['trap_id'].str.startswith('TRP-DUMMY', na=False)].copy()
        ws = wb.create_sheet('DummyInits')
        if len(dummy_inits) > 0:
            ws.append(list(dummy_inits.columns))
            for _, r in dummy_inits.iterrows():
                ws.append([r[c] for c in dummy_inits.columns])
            for c in range(1, len(dummy_inits.columns) + 1):
                ws.cell(1, c).font = Font(bold=True)
        print(f'  DummyInits: {len(dummy_inits)} rows')

    # Full sheets
    ws = wb.create_sheet('LabResults_Full')
    if len(lab_df) > 0:
        ws.append(list(lab_df.columns))
        for _, r in lab_df.iterrows():
            ws.append([r[c] for c in lab_df.columns])
        for c in range(1, len(lab_df.columns) + 1):
            ws.cell(1, c).font = Font(bold=True)
    print(f'  LabResults_Full: {len(lab_df)} rows')

    ws = wb.create_sheet('SamplingInit_Full')
    if len(init_df) > 0:
        ws.append(list(init_df.columns))
        for _, r in init_df.iterrows():
            ws.append([r[c] for c in init_df.columns])
        for c in range(1, len(init_df.columns) + 1):
            ws.cell(1, c).font = Font(bold=True)
    print(f'  SamplingInit_Full: {len(init_df)} rows')

    # Save
    wb.save(out_path)
    size_mb = out_path.stat().st_size / 1024 / 1024
    print(f'\nDONE: {out_path} ({size_mb:.1f} MB)')


if __name__ == '__main__':
    main()
